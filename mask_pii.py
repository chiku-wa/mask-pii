"""
mask_pii.py  --  日本語個人情報マスキングスクリプト
=====================================================
対象: 氏名 / 住所 / 部門名
方式: 3層アーキテクチャ
  Layer 1: 全角/半角正規化
  Layer 2: ルールベース（正規表現）
  Layer 3: GiNZA NER（spaCy日本語モデル）

セットアップ:
  python -m pip install openpyxl==3.1.5 click==8.1.8 spacy==3.7.5 ginza==5.2.0 ja-ginza==5.2.0 ja-ginza-electra==5.2.0

使い方:
  # テキスト直接指定
  python mask_pii.py --text "田中太郎（営業部）東京都渋谷区1-2-3"

  # テキストファイル入力
  python mask_pii.py --file input.txt --output result.json

  # Excelファイル入力 → マスク済みExcel出力
  python mask_pii.py --excel inquiries.xlsx --excel-out masked.xlsx

  # 部門名CSV併用
  python mask_pii.py --excel inquiries.xlsx --excel-out masked.xlsx --dept-csv 部門一覧.csv
"""

import re
import json
import csv
import argparse
import unicodedata
from dataclasses import dataclass, field, asdict

PREFECTURE_PATTERN = (
    r"北海道|東京都|(?:大阪|京都)府|(?:神奈川|埼玉|千葉|愛知|福岡|静岡|茨城|栃木|群馬|"
    r"新潟|富山|石川|福井|山梨|長野|岐阜|三重|滋賀|兵庫|奈良|和歌山|鳥取|島根|岡山|広島|"
    r"山口|徳島|香川|愛媛|高知|佐賀|長崎|熊本|大分|宮崎|鹿児島|沖縄|"
    r"青森|岩手|宮城|秋田|山形|福島)県"
)

DEPT_EXCLUDE_TERMS = {
    # 方向・範囲・関係を表す一般語。部門名CSVに入っていてもマスクしない。
    "外部", "内部", "上部", "下部", "中部", "前部", "後部",
    "一部", "全部", "各部", "両部", "大部", "半部",
    "他部", "当部", "同部", "該部", "本部", "支部",
    "社内", "社外", "全社", "自社", "他社", "弊社", "御社",
    "国内", "海外", "全体", "一式",
    # 組織・役割の一般名詞。単独では部門名として扱わない。
    "部", "部門", "部署", "課", "室", "係", "班",
    "チーム", "グループ", "センター", "組織", "会社", "企業", "法人",
    "担当", "担当者", "関係者", "責任者", "管理者",
    "社員", "職員", "役員", "利用者", "ユーザー", "顧客", "お客様",
}


# ---------------------------------------------------------------------------
# データクラス
# ---------------------------------------------------------------------------
@dataclass
class MaskRecord:
    category: str
    original: str
    replacement: str
    start: int
    end: int
    layer: str  # "rule" or "ner"

@dataclass
class MaskResult:
    masked_text: str
    mask_count: int
    records: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Layer 1: テキスト正規化（全角→半角など）
# ---------------------------------------------------------------------------
def normalize(text: str) -> str:
    return unicodedata.normalize("NFKC", text)


# ---------------------------------------------------------------------------
# Layer 2: ルールベースマスキング
# ---------------------------------------------------------------------------
RULE_PATTERNS = [
    ("電話番号",   r"0\d{1,4}[-－]\d{2,4}[-－]\d{4}", "[電話番号]"),
    ("メール",     r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", "[メール]"),
    ("郵便番号",   r"〒?\d{3}[-－]\d{4}", "[郵便番号]"),
    ("生年月日",   r"(昭和|平成|令和)\s*\d{1,2}年\s*\d{1,2}月\s*\d{1,2}日", "[生年月日]"),
    ("生年月日",   r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", "[生年月日]"),
    ("カード番号", r"\b\d{4}[\s\-]\d{4}[\s\-]\d{4}[\s\-]\d{4}\b", "[カード番号]"),
    ("住所",
     rf"({PREFECTURE_PATTERN})[^\s　、。]{{2,60}}?[0-9０-９\-－番地号室]+",
     "[住所]"),
    ("住所",
     rf"({PREFECTURE_PATTERN})[^\s　、。]{{1,20}}?[市区町村]",
     "[住所]"),
]

DEPT_NAME_PATTERN = re.compile(
    r"([A-Za-zＡ-Ｚａ-ｚ0-9０-９一-龥ァ-ヶー]{1,30}"
    r"(?:部門|事業部|本部|部署|課|室|係|チーム|グループ|センター)|"
    r"[A-Za-zＡ-Ｚａ-ｚ0-9０-９一-龥ァ-ヶー]{1,29}"
    r"(?<![一二三四五六七八九十数全各両大半外内上下中前後他当同該本支])部)"
    r"(?=の|に|へ|で|から|と|や|および|及び|ならびに|並びに|所属|です|となります|[,、。．.）)\]】」』]|$)"
)

DEPT_GENERIC_PART_PATTERN = re.compile(
    r"^[一二三四五六七八九十数全各両大半外内上下中前後他当同該本支].{0,2}部$"
)

DEPT_NER_LABELS = {
    "Department",
    "Division",
    "Section",
}

def apply_rules(text: str):
    records = []
    for category, pattern, label in RULE_PATTERNS:
        for m in sorted(re.finditer(pattern, text), key=lambda x: x.start(), reverse=True):
            records.append(MaskRecord(
                category=category, original=m.group(), replacement=label,
                start=m.start(), end=m.end(), layer="rule",
            ))
            text = text[:m.start()] + label + text[m.end():]
    return text, records


# ---------------------------------------------------------------------------
# 部門名パターン照合
# ---------------------------------------------------------------------------
def is_valid_dept_name(dept: str) -> bool:
    dept = normalize(dept.strip())
    if not dept or dept in DEPT_EXCLUDE_TERMS:
        return False
    return not DEPT_GENERIC_PART_PATTERN.fullmatch(dept)

def apply_dept_patterns(text: str):
    records = []
    for m in sorted(DEPT_NAME_PATTERN.finditer(text), key=lambda x: x.start(1), reverse=True):
        original = m.group(1)
        if not is_valid_dept_name(original):
            continue
        records.append(MaskRecord(
            category="部門名", original=original, replacement="[部門名]",
            start=m.start(1), end=m.end(1), layer="rule",
        ))
        text = text[:m.start(1)] + "[部門名]" + text[m.end(1):]
    return text, records


# ---------------------------------------------------------------------------
# 部門名辞書照合
# ---------------------------------------------------------------------------
def apply_dept_dict(text: str, dept_list: list):
    records = []
    valid_depts = [dept for dept in dept_list if is_valid_dept_name(dept)]
    for dept in sorted(valid_depts, key=len, reverse=True):
        if dept in text:
            records.append(MaskRecord(
                category="部門名", original=dept, replacement="[部門名]",
                start=-1, end=-1, layer="rule",
            ))
            text = text.replace(dept, "[部門名]")
    return text, records


# ---------------------------------------------------------------------------
# Layer 3: GiNZA NER
# ---------------------------------------------------------------------------
NER_LABEL_MAP = {
    "Person":       ("氏名",   "[氏名]"),
    "GPE":          ("住所",   "[住所]"),
    "Location":     ("住所",   "[住所]"),
    "Province":     ("住所",   "[住所]"),
    "City":         ("住所",   "[住所]"),
    "County":       ("住所",   "[住所]"),
    "Ward":         ("住所",   "[住所]"),
    "Station":      ("住所",   "[住所]"),
    "Facility":     ("施設名", "[施設名]"),
    "Organization": ("組織名", "[組織名]"),
}

def apply_ner(text: str, nlp):
    doc = nlp(text)
    records = []
    for ent in sorted(doc.ents, key=lambda e: e.start_char, reverse=True):
        if ent.label_ in DEPT_NER_LABELS and is_valid_dept_name(ent.text):
            label_info = ("部門名", "[部門名]")
        else:
            label_info = NER_LABEL_MAP.get(ent.label_)
        if not label_info:
            continue
        category, replacement = label_info
        records.append(MaskRecord(
            category=category, original=ent.text, replacement=replacement,
            start=ent.start_char, end=ent.end_char, layer="ner",
        ))
        text = text[:ent.start_char] + replacement + text[ent.end_char:]
    return text, records


# ---------------------------------------------------------------------------
# マスキング本体
# ---------------------------------------------------------------------------
def mask(text: str, dept_list: list, nlp) -> MaskResult:
    all_records = []
    text = normalize(text)
    text, r = apply_rules(text);         all_records.extend(r)
    text, r = apply_dept_patterns(text); all_records.extend(r)
    if dept_list:
        text, r = apply_dept_dict(text, dept_list); all_records.extend(r)
    if nlp:
        text, r = apply_ner(text, nlp);  all_records.extend(r)
    return MaskResult(masked_text=text, mask_count=len(all_records), records=all_records)


# ---------------------------------------------------------------------------
# Excel 入出力
# ---------------------------------------------------------------------------
def read_excel_column_a(path: str) -> list:
    """
    A列の全セルを読み取る。
    セル内改行（ALT+Enter）も含めてそのまま取得する。
    戻り値: [(行番号, テキスト), ...]  ※ヘッダー行も含む
    """
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        val = cell.value
        if val is None:
            rows.append((cell.row, ""))
        else:
            rows.append((cell.row, str(val)))
    return rows

def write_excel_masked(src_path: str, dst_path: str, masked_rows: list, quiet: bool = False):
    """
    元Excelを土台に、B列にマスク済みテキスト、C列にマスク件数を書き込んで保存。
    masked_rows: [(行番号, MaskResult), ...]
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font

    wb = load_workbook(src_path)
    ws = wb.active

    # ヘッダー色（1行目がヘッダーの場合に装飾）
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_font = Font(bold=True)

    # B1, C1 にヘッダーを設定
    ws["B1"] = "マスク済みテキスト"
    ws["C1"] = "マスク件数"
    for col in ("B1", "C1"):
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(wrap_text=True, vertical="top")

    for row_no, result in masked_rows:
        if row_no == 1:
            continue  # ヘッダー行はスキップ
        b_cell = ws.cell(row=row_no, column=2)
        c_cell = ws.cell(row=row_no, column=3)
        b_cell.value = result.masked_text
        b_cell.alignment = Alignment(wrap_text=True, vertical="top")
        c_cell.value = result.mask_count
        c_cell.alignment = Alignment(horizontal="center", vertical="top")

    # 列幅調整
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12

    wb.save(dst_path)
    if not quiet:
        print(f"マスク済みExcelを保存しました: {dst_path}")


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------
def load_dept_csv(path: str) -> list:
    depts = []
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if row:
                val = normalize(row[0].strip())
                if is_valid_dept_name(val):
                    depts.append(val)
    return depts

def load_nlp(quiet: bool = False):
    try:
        import spacy
        for model_name in ("ja_ginza_electra", "ja_core_news_trf", "ja_ginza"):
            try:
                nlp = spacy.load(model_name)
                if not quiet:
                    print(f"[NER] モデル読み込み: {model_name}")
                return nlp
            except OSError:
                continue
        print("[警告] GiNZA/spaCy 日本語モデルが見つかりません。ルールベースのみで処理します。")
        print("       python -m pip install ginza==5.2.0 ja-ginza==5.2.0 ja-ginza-electra==5.2.0  を実行してください。")
    except ImportError:
        print("[警告] spaCy がインストールされていません。ルールベースのみで処理します。")
    return None


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="日本語個人情報マスキングツール")
    # テキスト系
    parser.add_argument("--text",      help="マスキング対象テキスト（直接指定）")
    parser.add_argument("--file",      help="マスキング対象テキストファイル")
    parser.add_argument("--output",    help="テキスト/JSON結果の保存先")
    # Excel系
    parser.add_argument("--excel",     help="入力Excelファイル（A列を読み取る）")
    parser.add_argument("--excel-out", help="マスク済みExcel出力先（省略時は masked_<入力名>.xlsx）")
    # 共通
    parser.add_argument("--dept-csv",  help="部門名CSVファイルパス（省略可）")
    parser.add_argument("--no-ner",    action="store_true", help="NERを無効化")
    parser.add_argument("--quiet",     action="store_true", help="標準出力を抑制")
    args = parser.parse_args()

    dept_list = load_dept_csv(args.dept_csv) if args.dept_csv else []
    nlp = None if args.no_ner else load_nlp(args.quiet)

    # ── Excel モード ──────────────────────────────────────────
    if args.excel:
        import os
        rows = read_excel_column_a(args.excel)
        masked_rows = []
        total_masks = 0

        if not args.quiet:
            print(f"\n{len(rows)} 行を処理します...")

        for row_no, text in rows:
            if row_no == 1 or not text.strip():
                masked_rows.append((row_no, MaskResult(masked_text=text, mask_count=0)))
                continue
            result = mask(text, dept_list, nlp)
            masked_rows.append((row_no, result))
            total_masks += result.mask_count
            if not args.quiet:
                print(f"  行 {row_no:3d}: {result.mask_count} 件マスク")

        out_path = args.excel_out or (
            os.path.join(os.path.dirname(args.excel),
                         "masked_" + os.path.basename(args.excel))
        )
        write_excel_masked(args.excel, out_path, masked_rows, args.quiet)

        if not args.quiet:
            print(f"\n完了: 合計 {total_masks} 件をマスクしました。")
        return

    # ── テキスト / ファイル モード ────────────────────────────
    if args.text:
        input_text = args.text
    elif args.file:
        with open(args.file, encoding="utf-8") as f:
            input_text = f.read()
    else:
        parser.error("--text / --file / --excel のいずれかを指定してください")

    result = mask(input_text, dept_list, nlp)

    if not args.quiet:
        print("\n===== マスキング結果 =====")
        print(result.masked_text)
        print(f"\n===== マスクログ ({result.mask_count} 件) =====")
        if result.records:
            for r in result.records:
                print(f"  [{r.layer:4s}] [{r.category}]  {r.original!r}  →  {r.replacement}")
        else:
            print("  (マスク対象なし)")

    if args.output:
        out = {
            "masked_text": result.masked_text,
            "mask_count":  result.mask_count,
            "records": [asdict(r) for r in result.records],
        }
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        if not args.quiet:
            print(f"\n結果を保存しました: {args.output}")

if __name__ == "__main__":
    main()
